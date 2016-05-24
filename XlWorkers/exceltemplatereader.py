#!/usr/bin/python

__author__ = "Jackson"

from openpyxl import load_workbook
from XlWorkers.Config import Config
import XlWorkers.Utils as Utils


'''
Excel file reader
'''

class FOExcelReader(object):
    def __init__(self, filename,logger):
        # super.__init__(self)
        # resource identify as key value (language code as key value is translate) {"identify":{"en":"identify_en"}}
        self.mappingValue = {}

        self.logger = logger

        wb = load_workbook(filename, read_only=True)

        ws = wb.active
        self.logger("load excel to memory :" , filename)
        for content_row in ws.iter_rows(row_offset=Config.CONTENT_ROW_START - 1):
            #string resource identify cell objects
            key_cell = content_row[Config.KEY_COL_POS - 1]

            res_key = key_cell.value

            if Utils.is_str_valid(res_key):
            	
            	self.mappingValue[res_key] = {}

            	for cell_i in range(Config.LANG_COL_START - 1, len(content_row)):
                    #get translate values
                    cell = content_row[cell_i]
                    #mapping display language to language code
                    lang_col_info = Config.LANG_COL_POS_MAP.get(cell.column)
                    if lang_col_info is not None:
                    	self.mappingValue[res_key][lang_col_info.code] = cell.value

                    pass

        self.logger("will be translate keys:",self.mappingValue.keys())
        self.logger("completion load excel to memory")

    def hasResourceKey(self,res_key):
    	# print("res_key:",res_key)
    	res_values = None
    	try:
    		res_values = self.mappingValue[res_key]
    	except Exception as e:
    		print(e)
    	
    	if(res_values is not None):
    		return True
    	else:
    		return False

    def get_res_str(self, lang_code, res_key):
        res_values = self.mappingValue.get(res_key)
        if res_values is not None:
            return res_values.get(lang_code)
        return None
